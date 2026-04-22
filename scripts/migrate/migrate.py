"""
Jus Chiri — Migrador de imágenes v3
Motor: Playwright + Bing Images

Problemas resueltos vs v2:
  - DuckDuckGo daba 403 ratelimit a partir del producto ~10
  - Queries tenían "sneaker" y duplicaban la marca en el nombre
  - Sin deduplicación: 1,505 búsquedas para ~400 productos únicos
  - Sin checkpoint: si se interrumpía, empezaba desde cero

Soluciones:
  1. Playwright + Bing Images → da URLs reales, más tolerante al scraping
  2. Deduplicación por (marca, nombre_base) → busca 1 vez, reutiliza para variantes
  3. Checkpoint automático → resume desde el último producto procesado
  4. Backoff exponencial ante errores (30s → 60s → 120s)
  5. Queries limpios: sin "sneaker", sin duplicar marca

Columnas del sheet (orden):
  1. IMAGEN (ignorada) | 2. Stock anterior | 3. STOCK ACTUAL |
  4. Código | 5. Nombre | 6. Categorias | 7. Marca | 8. PRECIO

Uso:
  pip install -r requirements.txt
  playwright install chromium
  # Crea el .env en la raíz del proyecto con las variables necesarias
  python migrate.py
  # Para reanudar tras interrupción: python migrate.py  (usa checkpoint.json)
"""

import os
import re
import sys
import json
import csv
import time
import random
import logging
import requests
import openpyxl
import boto3
from io import BytesIO
from pathlib import Path
from dataclasses import dataclass, asdict
from typing import Optional
from dotenv import load_dotenv
from PIL import Image as PilImage

# Carga siempre el .env de la raíz del monorepo
load_dotenv(Path(__file__).resolve().parents[2] / ".env")

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# ── Configuración ─────────────────────────────────────────────────────────────

SHEET_ID    = os.environ.get("GOOGLE_SHEET_ID", "")
S3_BUCKET   = os.environ.get("S3_BUCKET2", "")
S3_REGION   = os.environ.get("S3_REGION2", "us-east-1")
S3_FOLDER   = os.environ.get("S3_FOLDER", "products")
API_URL     = os.environ.get("API_URL", "")

OUTPUT_DIR      = Path("output")
OUTPUT_JSON     = OUTPUT_DIR / "products.json"
OUTPUT_LOG      = OUTPUT_DIR / "migration_log.csv"
CHECKPOINT_FILE = OUTPUT_DIR / "checkpoint.json"

MIN_WIDTH   = 300
MIN_HEIGHT  = 300
SCORE_HIGH  = 70
SCORE_MEDIUM = 40

# Delay entre búsquedas (segundos) — más lento = menos bloqueos
DELAY_MIN   = 6.0
DELAY_MAX   = 12.0

# Dominios de confianza + bonus de score
TRUSTED_DOMAINS = {
    "nike.com": 40, "jordan.com": 40, "adidas.com": 40,
    "newbalance.com": 35, "supreme.com": 40, "palmangels.com": 40,
    "yeezy.com": 35, "stockx.com": 35, "goat.com": 35,
    "farfetch.com": 30, "ssense.com": 30, "mrporter.com": 30,
    "net-a-porter.com": 30, "endclothing.com": 25, "flightclub.com": 25,
    "stadiumgoods.com": 25, "kicksonfire.com": 20, "sneakernews.com": 20,
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}


# ── Modelos ───────────────────────────────────────────────────────────────────

@dataclass
class Candidate:
    url: str
    page_url: str = ""
    alt: str = ""
    width: int = 0
    height: int = 0
    source_domain: str = ""
    score: int = 0


@dataclass
class ProductResult:
    codigo: str
    nombre: str
    marca: str
    imagen_url: Optional[str] = None
    confidence: str = "NONE"
    final_score: int = 0
    candidate_url: str = ""
    reused_from: str = ""   # código base si fue deduplicado
    error: str = ""


# ── Limpieza de nombres ────────────────────────────────────────────────────────

def get_base_name(nombre: str) -> str:
    """
    Extrae el nombre base eliminando variantes de talla y color.
    "DOUBLED LOGO TEE - Talla M - Color Negro" → "DOUBLED LOGO TEE"
    """
    s = nombre
    s = re.sub(r'\s*-\s*Talla\s+\w+', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s*-\s*Color\s+.+$', '', s, flags=re.IGNORECASE)
    return s.strip()


def build_queries(product: dict) -> list[str]:
    """
    Dos queries limpias para Bing Images.
    - Sin "sneaker" (el catálogo incluye ropa, accesorios, etc.)
    - Sin duplicar la marca si ya está en el nombre
    """
    marca  = product["marca"].strip()
    nombre = get_base_name(product["nombre"])

    # Evita duplicar si la marca ya está al inicio del nombre
    if nombre.upper().startswith(marca.upper()):
        q1 = nombre
    else:
        q1 = f"{marca} {nombre}"

    q2 = f"{q1} official product"

    return [q1, q2]


def domain_of(url: str) -> str:
    try:
        from urllib.parse import urlparse
        return urlparse(url).netloc.lower().replace("www.", "")
    except Exception:
        return ""


# ── Checkpoint ─────────────────────────────────────────────────────────────────

def load_checkpoint() -> dict:
    """Carga resultados previos. Clave: codigo."""
    if CHECKPOINT_FILE.exists():
        try:
            return json.loads(CHECKPOINT_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_checkpoint(checkpoint: dict):
    CHECKPOINT_FILE.write_text(
        json.dumps(checkpoint, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


# ── Bing Images via Playwright ────────────────────────────────────────────────

def search_bing_images(page, query: str, n: int = 6) -> list[Candidate]:
    """
    Scrapea Bing Images con Playwright.
    El atributo `m` de cada resultado contiene JSON con:
      murl → URL original de la imagen
      purl → URL de la página fuente
      desc → descripción
      w, h → dimensiones
    """
    from urllib.parse import quote
    url = f"https://www.bing.com/images/search?q={quote(query)}&qft=+filterui:imagesize-large&form=IRFLTR"

    candidates = []
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=25_000)
        page.wait_for_timeout(random.uniform(1500, 2500))

        items = page.query_selector_all("a.iusc")
        for item in items[:n]:
            try:
                m_raw = item.get_attribute("m") or ""
                if not m_raw:
                    continue
                data = json.loads(m_raw)
                img_url  = data.get("murl", "")
                page_url = data.get("purl", "")
                desc     = data.get("desc", "")
                w        = int(data.get("w") or 0)
                h        = int(data.get("h") or 0)

                if img_url and img_url.startswith("http"):
                    candidates.append(Candidate(
                        url=img_url,
                        page_url=page_url,
                        alt=desc,
                        width=w,
                        height=h,
                        source_domain=domain_of(page_url),
                    ))
            except Exception:
                continue

    except Exception as e:
        log.warning(f"  Bing error: {e}")

    return candidates


# ── Scoring ───────────────────────────────────────────────────────────────────

def score_candidates(candidates: list[Candidate], product: dict) -> list[Candidate]:
    url_counts: dict[str, int] = {}
    for c in candidates:
        url_counts[c.url] = url_counts.get(c.url, 0) + 1

    nombre_base = get_base_name(product["nombre"]).lower()
    marca       = product["marca"].lower()
    codigo      = product["codigo"].lower()

    for c in candidates:
        s = 0
        url_l  = c.url.lower()
        alt_l  = c.alt.lower()
        page_l = c.page_url.lower()

        # SKU en URL o alt (señal muy fuerte)
        if codigo in url_l:   s += 55
        if codigo in alt_l:   s += 40
        if codigo in page_l:  s += 25

        # Marca en dominio o alt
        if marca in c.source_domain: s += 25
        if marca in alt_l:            s += 10

        # Palabras del nombre base en alt
        words = [w for w in nombre_base.split() if len(w) > 3]
        s += sum(6 for w in words if w in alt_l or w in url_l)

        # Resolución
        if c.width >= 800 and c.height >= 800:   s += 20
        elif c.width >= 400 and c.height >= 400: s += 10
        elif c.width >= 200 and c.height >= 200: s += 4

        # Penalizar imágenes muy pequeñas
        if c.width and c.height and (c.width < MIN_WIDTH or c.height < MIN_HEIGHT):
            s -= 25

        # Dominio de confianza
        for dom, bonus in TRUSTED_DOMAINS.items():
            if dom in c.source_domain:
                s += bonus
                break

        # Misma URL en varias búsquedas
        appearances = url_counts.get(c.url, 1)
        if appearances >= 2: s += 30

        c.score = max(s, 0)

    return sorted(candidates, key=lambda x: x.score, reverse=True)


# ── Validación de imagen descargada ──────────────────────────────────────────

def download_and_validate(url: str) -> Optional[bytes]:
    """Descarga y valida: formato, tamaño mínimo, aspect ratio."""
    try:
        r = requests.get(url, headers=HEADERS, timeout=15, stream=True)
        r.raise_for_status()
        data = r.content
        if len(data) < 5_000:
            return None
        img = PilImage.open(BytesIO(data))
        w, h = img.size
        if w < MIN_WIDTH or h < MIN_HEIGHT:
            return None
        if not (0.35 < w / h < 2.8):   # descarta banners/logos alargados
            return None
        return data
    except Exception:
        return None


def get_image_extension(data: bytes) -> str:
    if data[:8] == b"\x89PNG\r\n\x1a\n": return "png"
    if data[:3] == b"\xff\xd8\xff":       return "jpg"
    if data[:6] in (b"GIF87a", b"GIF89a"): return "gif"
    if data[:4] == b"RIFF" and data[8:12] == b"WEBP": return "webp"
    return "jpg"


def content_type_for(ext: str) -> str:
    return {"png": "image/png", "jpg": "image/jpeg",
            "gif": "image/gif", "webp": "image/webp"}.get(ext, "image/jpeg")


# ── S3 ────────────────────────────────────────────────────────────────────────

def upload_to_s3(s3_client, data: bytes, codigo: str, ext: str) -> str:
    key = f"{S3_FOLDER}/{codigo}.{ext}"
    s3_client.put_object(
        Bucket=S3_BUCKET,
        Key=key,
        Body=data,
        ContentType=content_type_for(ext),
    )
    return f"https://{S3_BUCKET}.s3.{S3_REGION}.amazonaws.com/{key}"


# ── Google Sheet ──────────────────────────────────────────────────────────────

def download_sheet(sheet_id: str) -> BytesIO:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    log.info("Descargando sheet...")
    for attempt in range(4):
        try:
            r = requests.get(url, timeout=90, stream=True, headers=HEADERS)
            r.raise_for_status()
            chunks = []
            for chunk in r.iter_content(chunk_size=16_384):
                if chunk:
                    chunks.append(chunk)
            return BytesIO(b"".join(chunks))
        except Exception as e:
            if attempt == 3:
                raise
            wait = (attempt + 1) * 5
            log.warning(f"  Reintento {attempt + 1}/3 en {wait}s: {e}")
            time.sleep(wait)


def parse_number(raw, cast=float):
    try:
        return cast(raw) if raw is not None else 0
    except (ValueError, TypeError):
        return 0


def parse_categorias(raw) -> list:
    if not raw:
        return []
    return [c.strip() for c in str(raw).split(",") if c.strip()]


def read_products_from_sheet(xlsx: BytesIO) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active
    log.info(f'Sheet: "{ws.title}" — {ws.max_row - 1} productos')
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cols = list(row) + [None] * 8
        _img, stock_ant, stock_act, codigo, nombre, cats, marca, precio = cols[:8]
        if not codigo:
            continue
        products.append({
            "codigo":         str(codigo).strip(),
            "nombre":         str(nombre or "").strip(),
            "stock_anterior": parse_number(stock_ant, int),
            "stock_actual":   parse_number(stock_act, int),
            "categorias":     parse_categorias(cats),
            "marca":          str(marca or "").strip(),
            "precio":         parse_number(precio, float),
        })
    return products


# ── Búsqueda + subida para un producto ────────────────────────────────────────

def process_product(product: dict, page, s3_client) -> ProductResult:
    result = ProductResult(
        codigo=product["codigo"],
        nombre=product["nombre"],
        marca=product["marca"],
    )

    queries = build_queries(product)
    all_candidates: list[Candidate] = []

    for i, query in enumerate(queries):
        log.info(f"  [{i+1}/{len(queries)}] Bing: «{query}»")
        hits = search_bing_images(page, query, n=6)
        all_candidates.extend(hits)
        if i < len(queries) - 1:
            time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    if not all_candidates:
        result.error = "Sin candidatos"
        return result

    all_candidates = score_candidates(all_candidates, product)
    best = all_candidates[0]

    log.info(f"  Mejor candidato → score={best.score}  {best.url[:80]}")

    result.candidate_url = best.url
    result.final_score   = best.score
    result.confidence    = (
        "HIGH"   if best.score >= SCORE_HIGH   else
        "MEDIUM" if best.score >= SCORE_MEDIUM else
        "LOW"
    )

    if result.confidence == "LOW":
        log.warning(f"  Score bajo ({best.score}) — se sube igual, revisar en log")

    # Descargar y validar; si falla, intentar siguientes candidatos
    data = None
    for candidate in all_candidates[:5]:
        data = download_and_validate(candidate.url)
        if data:
            result.candidate_url = candidate.url
            result.final_score   = candidate.score
            break

    if not data:
        result.error = "Imagen no descargable o inválida"
        log.warning(f"  {result.error}")
        return result

    # Subir a S3
    ext = get_image_extension(data)
    try:
        result.imagen_url = upload_to_s3(s3_client, data, product["codigo"], ext)
        log.info(f"  ✓ S3: {result.imagen_url}")
    except Exception as e:
        result.error = f"Error S3: {e}"
        log.error(f"  {result.error}")

    return result


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    if not SHEET_ID:  sys.exit("ERROR: GOOGLE_SHEET_ID no definido en .env")
    if not S3_BUCKET: sys.exit("ERROR: S3_BUCKET no definido en .env")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Leer sheet
    xlsx     = download_sheet(SHEET_ID)
    products = read_products_from_sheet(xlsx)
    log.info(f"{len(products)} productos en el sheet\n")

    # Cargar checkpoint (para reanudar si se interrumpió)
    checkpoint: dict[str, dict] = load_checkpoint()
    pending = [p for p in products if p["codigo"] not in checkpoint]
    log.info(f"Checkpoint: {len(checkpoint)} procesados, {len(pending)} pendientes\n")

    # Deduplicación: (marca, nombre_base) → resultado ya encontrado
    # Evita buscar 5 veces "DOUBLED LOGO TEE" por variantes de talla
    image_cache: dict[str, str] = {}   # clave: "marca|nombre_base" → imagen_url

    # Pre-cargar cache desde el checkpoint
    for r in checkpoint.values():
        if r.get("imagen_url") and not r.get("reused_from"):
            p_nombre = r.get("nombre", "")
            p_marca  = r.get("marca", "")
            cache_key = f"{p_marca}|{get_base_name(p_nombre)}"
            if cache_key not in image_cache:
                image_cache[cache_key] = r["imagen_url"]

    if not pending:
        log.info("Todos los productos ya procesados. Generando outputs...")
    else:
        s3 = boto3.client(
            "s3", region_name=S3_REGION,
            aws_access_key_id=os.environ.get("AWS_ACCESS_KEY_ID2"),
            aws_secret_access_key=os.environ.get("AWS_SECRET_ACCESS_KEY2"),
        )

        from playwright.sync_api import sync_playwright

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                user_agent=HEADERS["User-Agent"],
                locale="en-US",
                viewport={"width": 1280, "height": 900},
            )
            page = context.new_page()
            page.route("**/*.{woff,woff2,ttf,otf}", lambda r: r.abort())
            page.route("**/analytics*", lambda r: r.abort())

            total = len(pending)
            try:
                for i, product in enumerate(pending, 1):
                    codigo    = product["codigo"]
                    cache_key = f"{product['marca']}|{get_base_name(product['nombre'])}"

                    log.info(f"\n[{i}/{total}] {codigo} — {product['nombre']}")

                    if cache_key in image_cache:
                        log.info(f"  Reutilizando imagen de variante anterior")
                        result = ProductResult(
                            codigo=codigo,
                            nombre=product["nombre"],
                            marca=product["marca"],
                            imagen_url=image_cache[cache_key],
                            confidence="HIGH",
                            reused_from=cache_key,
                        )
                    else:
                        backoff = 30
                        for attempt in range(3):
                            try:
                                result = process_product(product, page, s3)
                                break
                            except Exception as e:
                                log.warning(f"  Error intento {attempt+1}: {e}. Esperando {backoff}s...")
                                time.sleep(backoff)
                                backoff *= 2
                        else:
                            result = ProductResult(
                                codigo=codigo, nombre=product["nombre"],
                                marca=product["marca"], error="Falló tras 3 intentos"
                            )

                        if result.imagen_url:
                            image_cache[cache_key] = result.imagen_url

                        if i < total:
                            delay = random.uniform(DELAY_MIN, DELAY_MAX)
                            log.info(f"  Esperando {delay:.1f}s...")
                            time.sleep(delay)

                    # Guardar checkpoint después de CADA producto
                    checkpoint[codigo] = asdict(result)
                    save_checkpoint(checkpoint)

            except KeyboardInterrupt:
                log.warning(
                    f"\n\nInterrumpido en producto {i}/{total} ({codigo})."
                    f"\nCheckpoint guardado → vuelve a ejecutar 'python migrate.py' para continuar."
                )
                save_checkpoint(checkpoint)
                browser.close()
                sys.exit(0)

            browser.close()

        save_checkpoint(checkpoint)

    # ── Generar outputs finales ───────────────────────────────────────────────
    results = [ProductResult(**checkpoint.get(p["codigo"], asdict(ProductResult(
        codigo=p["codigo"], nombre=p["nombre"], marca=p["marca"], error="No procesado"
    )))) for p in products]

    final_products = []
    for product, result in zip(products, results):
        final_products.append({**product, "imagen_url": result.imagen_url})

    OUTPUT_JSON.write_text(
        json.dumps(final_products, indent=2, ensure_ascii=False),
        encoding="utf-8"
    )

    with OUTPUT_LOG.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "codigo", "nombre", "marca", "confidence",
            "final_score", "imagen_url", "candidate_url", "reused_from", "error"
        ])
        w.writeheader()
        for r in results:
            w.writerow(asdict(r))

    # Resumen
    high   = sum(1 for r in results if r.confidence == "HIGH")
    medium = sum(1 for r in results if r.confidence == "MEDIUM")
    low    = sum(1 for r in results if r.confidence == "LOW")
    reused = sum(1 for r in results if r.reused_from)
    errors = sum(1 for r in results if r.error)

    log.info(
        f"\n── Resumen ─────────────────────────────────\n"
        f"  Total     : {len(results)}\n"
        f"  HIGH      : {high}\n"
        f"  MEDIUM    : {medium}\n"
        f"  LOW       : {low}  ← revisar migration_log.csv\n"
        f"  Reutilizados (variantes): {reused}\n"
        f"  Sin imagen  : {errors}\n"
        f"  Búsquedas reales realizadas: {high + medium + low - reused}\n"
    )

    # Importar al backend (opcional)
    if API_URL:
        log.info(f"\nImportando al backend: {API_URL}/api/products/bulk ...")
        try:
            r = requests.post(f"{API_URL}/api/products/bulk", json=final_products, timeout=60)
            r.raise_for_status()
            res = r.json()
            log.info(f"  ✓ upserted={res.get('upserted', 0)}, modified={res.get('modified', 0)}")
        except Exception as e:
            log.error(f"  ERROR: {e}")


if __name__ == "__main__":
    main()
